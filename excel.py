import csv
import os

import openpyxl
from xls2xlsx import XLS2XLSX


def xls_to_xlsx(dir_path):

    """This function iterates over the specified directory and converts each .XLS file into .XLSX"""
    found_xls_file = False

    for filename in os.listdir(dir_path):
        if filename.endswith(".xls"):
            found_xls_file = True
            xls_file_path = os.path.join(dir_path, filename)
            xlsx_file_path = os.path.join(dir_path, filename.replace(".xls", ".xlsx"))
            x2x = XLS2XLSX(xls_file_path)
            x2x.to_xlsx(xlsx_file_path)
            print(f"[INFO] File {xls_file_path} converted to .XLSX")

    if not found_xls_file:
        print(f"[WARN] There are no .xls files in {dir_path}")           


def excel_to_csv(dir_path):
    """
    This function converts every sheet of a .xlsx file in the specified directory into .csv file
    and stores it into a newly created directory with the name of the .xlsx file.
    """
    # Calling xls_to_xlsx func
    xls_to_xlsx(dir_path)

    found_xlsx_file = False

    for filename in os.listdir(dir_path):
        # Checks if file name ends with .XLSX
        if filename.endswith(".xlsx"):

            found_xlsx_file = True

            # Creates an absolute file path to the .XLSX file
            file_path = os.path.join(dir_path, filename)
            
            # Creates a new directory in the folder of the .XLSX file with its name(There it will store the .CSV files)
            new_dir_path = os.path.join(dir_path, filename.rstrip(".xlsx"))
            os.makedirs(new_dir_path, exist_ok=True)
            print(f"[INFO] Directory {new_dir_path} created.")
            
            # Opens the .XLSX workbook
            wb = openpyxl.load_workbook(file_path)

            # Loops over the each sheet in the workbook
            for sheet in wb.sheetnames:
                ws = wb[sheet]

                # Creates a list of all merged cells in the sheet
                merged_cells = list(ws.merged_cells)

                # Loops over each merged cell
                for mc in merged_cells:
                    # Unpacking the bounds of the merged cell
                    min_col, min_row, max_col, max_row = mc.bounds
                    # Getting the value of the top left cell
                    top_left_value = ws.cell(row=min_row, column=min_col).value
                    # Unmerging the cell
                    ws.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
                    # Loops over each row in the cell
                    for row in ws.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                        # Loops over each cell in the row and assigns it the value of the top left cell
                        for cell in row:
                            cell.value = top_left_value

                # Creates a list of lists representing all the data in the worksheet. 
                # Each inner list represents a row and contains the values of the cells in that row.
                data = [[cell.value for cell in row] for row in ws.iter_rows()]

                # Removing empty rows and columns
                data = [row for row in data if any(cell is not None and cell != "" for cell in row)] # Removes empty rows
                # Removes empty columns in reverse order checking if all the cell are None or empty
                for i in reversed(range(len(data[0]))): 
                    if all(row[i] is None or row[i] == "" for row in data):
                        for row in data:
                            del row[i]
                # Creates the path to the .CSV file
                path = os.path.join(new_dir_path, f"{sheet}.csv")
                # Opens a new CSV file in write mode
                with open(path, "w", newline="", encoding="utf-8-sig") as f:
                    # Creating a CSV writer object
                    writer = csv.writer(f)
                    # Writing the data into it
                    writer.writerows(data)
                    
                # Printing a message indicating that the file has been created
                print(f"[INFO] File {path} created")
                
    if not found_xlsx_file:
        print(f"[WARN] There are no .xlsx files in {dir_path}")


# ATTENTION

# SUPPORTED FEATURES BY xls2xlsx:
"""
- Multiple worksheets
- Text, Numbers, Dates/Times, Unicode
- Fonts, text color, bold, italic, underline, double underline, strikeout
- Solid and Pattern Fills with color
- Borders: Solid, Hair, Thin, Thick, Double, Dashed, Dotted; with color
- Alignment: Horizontal, Vertical, Rotated, Indent, Shrink To Fit
- Number Formats, including unicode currency symbols
- Hidden Rows and Columns
- Merged Cells
- Hyperlinks (only 1 per cell)
- Comments
"""

# NOT SUPPORTED FEATURES BY xls2xlsx:
"""
- Conditional Formatting (the current stylings are preserved)
- Formulas (the calculated values are preserved)
- Charts (the image of the chart is handled by .htm and .mht input formats)
- Drawings (the image of the drawing is handled by .htm and .mht input formats)
- Pivot tables (the current data is preserved)
- Text boxes (converted to an image by .htm and .mht input formats)
- Shapes and Clip Art (converted to an image by .htm and .mht input formats)
- Autofilter (the current filtered out rows are preserved)
- Rich text in cells (openpyxl does not support this: only styles applied to the entire cell are preserved)
- Named Ranges
- Macros (VBA)
"""